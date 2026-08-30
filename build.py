# -*- coding: utf-8 -*-
"""Сборка датасета дэшборда Альфа-КПД напрямую из Google-таблицы.
Источники: «Недельный отчёт по трафику» + «Месячный 2026»/«Месячный 2025».
Выручка = сумма УСПЕШНЫХ сделок. Сумма созданных лежит отдельно (pipeline).
Расход НЕ округляется (округление давало дрейф +0,2% на 84 неделях).
"""
import openpyxl, json, datetime, collections, sys

SRC='book2.xlsx'
MONNO={'Январь':1,'Февраль':2,'Март':3,'Апрель':4,'Май':5,'Июнь':6,
       'Июль':7,'Август':8,'Сентябрь':9,'Октябрь':10,'Ноябрь':11,'Декабрь':12}
MONRU={v:k for k,v in MONNO.items()}

def num(v):
    if v is None: return 0.0
    if isinstance(v,str):
        v=v.replace('\xa0','').replace(' ','').replace('\u2009','').replace(',','.').strip()
        if v in ('','-','—','–'): return 0.0
    try: return float(v)
    except: return 0.0

def txt(v):
    if v is None: return None
    v=str(v).replace('\xa0',' ').strip()
    return None if v in ('','-','—','–') else v

def iso(d):
    if isinstance(d,datetime.datetime): return d.date().isoformat()
    if isinstance(d,datetime.date): return d.isoformat()
    return None

wb=openpyxl.load_workbook(SRC,data_only=True,read_only=True)
issues=[]

# ---------- НЕДЕЛИ ----------
weekly=[]
for r in wb['Недельный отчёт по трафику'].iter_rows(min_row=2,values_only=True):
    if not r or r[0] is None or r[2] is None: continue
    st=iso(r[0])
    if not st: continue
    d=txt(r[2]); s=txt(r[3]); p=txt(r[4])
    row=dict(period=st, end=iso(r[1]), dir=d, src=s, page=p,
        spend=num(r[5]), leads=int(num(r[6])), kl=int(num(r[8])),
        inwork=int(num(r[9])), badlead=int(num(r[10])), spam=int(num(r[11])),
        existing=int(num(r[13])),
        deals_created=int(num(r[17])), pipeline=num(r[18]),
        deals_lost=int(num(r[19])), revenue_lost=num(r[20]),
        deals_won=int(num(r[21])), revenue_won=num(r[22]))
    if row['kl']>row['leads']:
        issues.append(dict(kind='kl-above-leads',period=st,dir=d,src=s,leads=row['leads'],kl=row['kl']))
    weekly.append(row)

# ---------- МЕСЯЦЫ ----------
monthly=[]
def g(r,i):
    return r[i] if i < len(r) else None

def read_month(tab, year):
    hdr_done=False
    for r in wb[tab].iter_rows(min_row=1,values_only=True):
        if not r or len(r)<15: continue
        if r[1]=='Месяц': hdr_done=True; continue
        if not hdr_done: continue
        mn=txt(r[1]); d=txt(r[2])
        if not mn or not d or mn not in MONNO: continue
        no=MONNO[mn]
        monthly.append(dict(period=f'{year}-{no:02d}', year=year, monthNo=no, monthName=mn,
            dir=d, src=txt(r[3]), page=txt(r[4]),
            spend=num(g(r,5)), leads=int(num(g(r,6))), kl=int(num(g(r,8))),
            badlead=int(num(g(r,9))), spam=int(num(g(r,10))), existing=int(num(g(r,12))),
            deals_created=int(num(g(r,16))), pipeline=num(g(r,17)),
            deals_lost=int(num(g(r,18))), revenue_lost=num(g(r,19)),
            deals_won=int(num(g(r,20))), revenue_won=num(g(r,21))))
read_month('Месячный 2026',2026)
try: read_month('Месячный 2025',2025)
except Exception as e: print('2025 skip',e,file=sys.stderr)

# ---------- ПЛАН ----------
old=json.load(open('v7_data.json'))
plan_targets=old.get('plan_targets',{})

# ФАКТ план-факта считаем ИЗ ТЕХ ЖЕ строк (закрывает блокер 4)
plan_fact={}
mf=collections.defaultdict(collections.Counter)
for r in monthly:
    k=(r['period'],r['dir'])
    for f in ('spend','leads','kl','deals_created','deals_won','revenue_won'): mf[k][f]+=r[f]
periods=sorted(set(r['period'] for r in monthly if r['year']==2026))
dirs_all=sorted(set(r['dir'] for r in monthly if r['year']==2026))
for per in periods:
    tg={t['dir']:t for t in plan_targets.get(per,[])}
    rows=[]
    for d in sorted(set(list(tg.keys())+[x for x in dirs_all if mf[(per,x)]['spend'] or mf[(per,x)]['kl']])):
        f=mf[(per,d)]; t=tg.get(d)
        rows.append(dict(dir=d,
            kl_plan=round(t['kl_plan'],1) if t else None,
            leads_plan=round(t['leads_plan'],1) if t else None,
            spend_plan=round(t['spend_plan_traffic'],2) if t else None,
            revenue_plan=round(t['revenue_plan'],2) if t else None,
            kl_fact=f['kl'], leads_fact=f['leads'], spend_fact=round(f['spend'],2),
            deals_fact=f['deals_won'], revenue_fact=f['revenue_won']))
    plan_fact[per]=rows

wcov=sorted(set(r['period'] for r in weekly))
mcov=sorted(set(r['period'] for r in monthly))
data=dict(
  meta=dict(builtAt=datetime.datetime.now(datetime.timezone.utc).isoformat(),
    source=dict(sheetId='1lIRkcFWVQMZLg_TNEGjNH_6FN7UXJfl62Yjo0xOqBus',
      tabs=['Недельный отчёт по трафику','Месячный 2026','Месячный 2025']),
    rules=['Выручка = сумма УСПЕШНЫХ сделок (кол. V). Сумма созданных = pipeline, это потенциал, не деньги.',
           'Расход не округляется.',
           'План-факт: факт считается из строк месячного листа, отдельной копии нет.',
           'Лист «янв-июнь 2026» НЕ используется: двойной счёт Q1 + перепутанные колонки.'],
    coverage=dict(weeks=len(wcov),weekFrom=wcov[0],weekTo=wcov[-1],
                  months=len(mcov),monthFrom=mcov[0],monthTo=mcov[-1]),
    issues=issues[:400], issuesTotal=len(issues)),
  weekly=weekly, monthly=monthly,
  plan_fact=plan_fact, plan_targets=plan_targets)
json.dump(data,open('data.json','w'),ensure_ascii=False,separators=(',',':'))

print('weekly',len(weekly),'monthly',len(monthly))
print('недель',len(wcov),wcov[0],'..',wcov[-1])
print('месяцев',len(mcov),mcov[0],'..',mcov[-1])
print('расход недельный %.2f'%sum(r['spend'] for r in weekly))
print('расход 2026 мес %.2f'%sum(r['spend'] for r in monthly if r['year']==2026))
print('выручка 2026 %.0f'%sum(r['revenue_won'] for r in monthly if r['year']==2026))
print('issues',len(issues))
